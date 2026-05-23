import os
from pathlib import Path

import openpyxl


SOURCE_FILE = "水映南瀛_客戶列表_20260520.xlsx"


def sql_quote(value):
    return "'" + str(value or "").replace("'", "''") + "'"


def sql_number(value):
    if value in (None, ""):
        return "NULL"
    try:
        return str(float(str(value).replace(",", "")))
    except Exception:
        return "NULL"


def main():
    src = os.environ["XLSX_PATH"]
    out = Path(os.environ["OUT_SQL"])
    out.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.load_workbook(src, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(value).strip() if value is not None else "" for value in next(sheet.iter_rows(values_only=True))]
    indexes = {header: index for index, header in enumerate(headers)}

    def cell(row, name):
        index = indexes.get(name)
        if index is None or index >= len(row):
            return ""
        value = row[index]
        if value is None:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d %H:%M:%S") if name == "領1期點時間" else value.strftime("%Y-%m-%d")
        return str(value).strip()

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        member_id = cell(row, "ID")
        if not member_id:
            continue
        rows.append({
            "member_id": member_id,
            "customer_name": cell(row, "會員姓名"),
            "gender": cell(row, "性別"),
            "birthday": cell(row, "生日"),
            "residence_area": cell(row, "居住地區"),
            "phone": cell(row, "手機"),
            "credit_balance": cell(row, "01購物金餘額"),
            "claimed_at": cell(row, "領1期點時間"),
        })

    lines = [
        "BEGIN TRANSACTION;",
        (
            "CREATE TABLE IF NOT EXISTS crm_customers ("
            "member_id TEXT PRIMARY KEY, "
            "line_user_id TEXT NOT NULL DEFAULT '', "
            "customer_name TEXT NOT NULL DEFAULT '', "
            "gender TEXT NOT NULL DEFAULT '', "
            "birthday TEXT NOT NULL DEFAULT '', "
            "residence_area TEXT NOT NULL DEFAULT '', "
            "phone TEXT NOT NULL DEFAULT '', "
            "credit_balance REAL, "
            "claimed_at TEXT NOT NULL DEFAULT '', "
            "source_file TEXT NOT NULL DEFAULT '', "
            "imported_at TEXT NOT NULL DEFAULT (datetime('now')), "
            "updated_at TEXT NOT NULL DEFAULT (datetime('now'))"
            ");"
        ),
        "CREATE INDEX IF NOT EXISTS idx_crm_customers_phone ON crm_customers(phone);",
        "CREATE INDEX IF NOT EXISTS idx_crm_customers_residence ON crm_customers(residence_area);",
        "CREATE INDEX IF NOT EXISTS idx_crm_customers_gender ON crm_customers(gender);",
        "CREATE INDEX IF NOT EXISTS idx_crm_customers_claimed ON crm_customers(claimed_at);",
    ]

    for row in rows:
        values = [
            sql_quote(row["member_id"]),
            sql_quote(row["customer_name"]),
            sql_quote(row["gender"]),
            sql_quote(row["birthday"]),
            sql_quote(row["residence_area"]),
            sql_quote(row["phone"]),
            sql_number(row["credit_balance"]),
            sql_quote(row["claimed_at"]),
            sql_quote(SOURCE_FILE),
        ]
        lines.append(
            "INSERT INTO crm_customers ("
            "member_id, customer_name, gender, birthday, residence_area, phone, "
            "credit_balance, claimed_at, source_file, imported_at, updated_at"
            ") VALUES ("
            + ",".join(values)
            + ", datetime('now'), datetime('now')) "
            "ON CONFLICT(member_id) DO UPDATE SET "
            "customer_name=excluded.customer_name, "
            "gender=excluded.gender, "
            "birthday=excluded.birthday, "
            "residence_area=excluded.residence_area, "
            "phone=excluded.phone, "
            "credit_balance=excluded.credit_balance, "
            "claimed_at=excluded.claimed_at, "
            "source_file=excluded.source_file, "
            "updated_at=excluded.updated_at;"
        )

    lines.append("COMMIT;")
    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"rows={len(rows)}")
    print(out)


if __name__ == "__main__":
    main()

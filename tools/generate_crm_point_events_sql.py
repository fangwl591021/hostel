import os
from pathlib import Path

import openpyxl


SOURCE_FILE = "aibot_user_point_tainan_20260520.xlsx"


def sql_quote(value):
    return "'" + str(value or "").replace("'", "''") + "'"


def sql_number(value):
    if value in (None, ""):
        return "NULL"
    try:
        return str(float(str(value).replace(",", "")))
    except Exception:
        return "NULL"


def main():
    src = os.environ["XLSX_PATH"]
    out = Path(os.environ["OUT_SQL"])
    out.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(src, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(value).strip() if value is not None else "" for value in next(sheet.iter_rows(values_only=True))]
    indexes = {header: index for index, header in enumerate(headers)}

    def cell(row, name):
        index = indexes.get(name)
        if index is None or index >= len(row):
            return ""
        value = row[index]
        if value is None:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return str(value).strip()

    lines = [
        (
            "CREATE TABLE IF NOT EXISTS crm_point_events (event_row_id TEXT PRIMARY KEY, uuid TEXT NOT NULL DEFAULT '', "
            "member_id TEXT NOT NULL DEFAULT '', shop_user_id TEXT NOT NULL DEFAULT '', shop_id TEXT NOT NULL DEFAULT '', "
            "event_id TEXT NOT NULL DEFAULT '', event_name TEXT NOT NULL DEFAULT '', event_content TEXT NOT NULL DEFAULT '', "
            "point_type TEXT NOT NULL DEFAULT '', get_point REAL, point_balance REAL, check_name TEXT NOT NULL DEFAULT '', "
            "check_people_id TEXT NOT NULL DEFAULT '', check_phone TEXT NOT NULL DEFAULT '', room_number TEXT NOT NULL DEFAULT '', "
            "sub_shop_name TEXT NOT NULL DEFAULT '', child_shop_attribute TEXT NOT NULL DEFAULT '', sub_shop_area TEXT NOT NULL DEFAULT '', "
            "child_shop_ip TEXT NOT NULL DEFAULT '', child_shop_remark TEXT NOT NULL DEFAULT '', event_period TEXT NOT NULL DEFAULT '', "
            "created_at TEXT NOT NULL DEFAULT '', out_at TEXT NOT NULL DEFAULT '', source_file TEXT NOT NULL DEFAULT '', "
            "imported_at TEXT NOT NULL DEFAULT (datetime('now')), updated_at TEXT NOT NULL DEFAULT (datetime('now')));"
        ),
        "CREATE INDEX IF NOT EXISTS idx_crm_point_events_member ON crm_point_events(member_id);",
        "CREATE INDEX IF NOT EXISTS idx_crm_point_events_event ON crm_point_events(event_name);",
        "CREATE INDEX IF NOT EXISTS idx_crm_point_events_shop ON crm_point_events(sub_shop_name);",
        "CREATE INDEX IF NOT EXISTS idx_crm_point_events_created ON crm_point_events(created_at);",
    ]
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_id = cell(row, "id")
        if not row_id:
            continue
        count += 1
        values = [
            sql_quote(row_id), sql_quote(cell(row, "uuid")), sql_quote(cell(row, "user_id")),
            sql_quote(cell(row, "shop_user_id")), sql_quote(cell(row, "shop_id")), sql_quote(cell(row, "event_id")),
            sql_quote(cell(row, "event_name")), sql_quote(cell(row, "event_content")), sql_quote(cell(row, "point_type")),
            sql_number(cell(row, "get_point")), sql_number(cell(row, "point_balance")),
            sql_quote(cell(row, "check_name")), sql_quote(cell(row, "check_people_id")), sql_quote(cell(row, "check_phone")),
            sql_quote(cell(row, "room_number")), sql_quote(cell(row, "sub_shop_name")), sql_quote(cell(row, "child_shop_attribute")),
            sql_quote(cell(row, "sub_shop_area")), sql_quote(cell(row, "child_shop_ip")), sql_quote(cell(row, "child_shop_remark")),
            sql_quote(cell(row, "even_period")), sql_quote(cell(row, "created_at")), sql_quote(cell(row, "out_at")),
            sql_quote(SOURCE_FILE),
        ]
        lines.append(
            "INSERT INTO crm_point_events (event_row_id, uuid, member_id, shop_user_id, shop_id, event_id, event_name, event_content, "
            "point_type, get_point, point_balance, check_name, check_people_id, check_phone, room_number, sub_shop_name, "
            "child_shop_attribute, sub_shop_area, child_shop_ip, child_shop_remark, event_period, created_at, out_at, source_file, imported_at, updated_at) "
            "VALUES (" + ",".join(values) + ", datetime('now'), datetime('now')) "
            "ON CONFLICT(event_row_id) DO UPDATE SET member_id=excluded.member_id, event_name=excluded.event_name, event_content=excluded.event_content, "
            "point_type=excluded.point_type, get_point=excluded.get_point, point_balance=excluded.point_balance, check_name=excluded.check_name, "
            "check_phone=excluded.check_phone, room_number=excluded.room_number, sub_shop_name=excluded.sub_shop_name, sub_shop_area=excluded.sub_shop_area, "
            "created_at=excluded.created_at, out_at=excluded.out_at, source_file=excluded.source_file, updated_at=excluded.updated_at;"
        )
    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"rows={count}")
    print(out)


if __name__ == "__main__":
    main()
